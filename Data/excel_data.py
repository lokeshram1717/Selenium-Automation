from openpyxl import load_workbook

def get_test_data_from_excel(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        data.append(dict(zip(headers,row)))

    return data





